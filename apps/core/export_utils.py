"""
Utilidad centralizada para exportar datos a Excel (.xlsx) con filtros de período.
Formato soportado: Trimestral, Semestral, Anual, Todo
"""

import io
from datetime import date, datetime
from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─── Color palette ────────────────────────────────────────────────────────────
HEADER_FILL_COLOR  = "1A5276"   # Dark navy-blue
HEADER_FONT_COLOR  = "FFFFFF"
ALT_ROW_FILL_COLOR = "EAF4FB"   # Light blue-grey for alternating rows
# NOTE: No emojis in cell values — openpyxl can corrupt .xlsx on some platforms.


# ─── Period helpers ───────────────────────────────────────────────────────────

def get_period_range(periodo: str, anio: int | None = None):
    """
    Returns (date_from, date_to) for the given period string.
    periodo: 'trimestre1' .. 'trimestre4', 'semestre1', 'semestre2', 'anual', 'todo'
    Returns None for 'todo'.
    """
    year = anio or timezone.now().year
    ranges = {
        'trimestre1': (date(year, 1, 1),  date(year, 3, 31)),
        'trimestre2': (date(year, 4, 1),  date(year, 6, 30)),
        'trimestre3': (date(year, 7, 1),  date(year, 9, 30)),
        'trimestre4': (date(year, 10, 1), date(year, 12, 31)),
        'semestre1':  (date(year, 1, 1),  date(year, 6, 30)),
        'semestre2':  (date(year, 7, 1),  date(year, 12, 31)),
        'anual':      (date(year, 1, 1),  date(year, 12, 31)),
    }
    return ranges.get(periodo)


def get_period_label(periodo: str, anio: int | None = None) -> str:
    year = anio or timezone.now().year
    labels = {
        'trimestre1': f"T1 {year} (Ene-Mar)",
        'trimestre2': f"T2 {year} (Abr-Jun)",
        'trimestre3': f"T3 {year} (Jul-Sep)",
        'trimestre4': f"T4 {year} (Oct-Dic)",
        'semestre1':  f"1er Semestre {year} (Ene-Jun)",
        'semestre2':  f"2do Semestre {year} (Jul-Dic)",
        'anual':      f"Anio {year} (Completo)",
        'todo':       "Historial Completo",
    }
    return labels.get(periodo, "Historial")


# ─── Internal helpers ─────────────────────────────────────────────────────────

def _make_styles():
    """Returns a dict of reusable openpyxl style objects."""
    return dict(
        header_fill   = PatternFill("solid", fgColor=HEADER_FILL_COLOR),
        alt_fill      = PatternFill("solid", fgColor=ALT_ROW_FILL_COLOR),
        # PatternFill(fill_type=None) is the only safe way to have 'no fill'
        # without generating malformed XML in openpyxl.
        no_fill       = PatternFill(fill_type=None),
        header_font   = Font(bold=True, color=HEADER_FONT_COLOR, size=11),
        title_font    = Font(bold=True, size=14),
        subtitle_font = Font(italic=True, size=10, color="555555"),
        center        = Alignment(horizontal="center", vertical="center"),
        left_vc       = Alignment(vertical="center"),
        border        = Border(
            left   = Side(style="thin", color="CCCCCC"),
            right  = Side(style="thin", color="CCCCCC"),
            top    = Side(style="thin", color="CCCCCC"),
            bottom = Side(style="thin", color="CCCCCC"),
        ),
    )


def _write_sheet(ws, headers: list, rows: list, title: str, period_label: str, styles: dict):
    """
    Writes a fully styled sheet into an existing openpyxl Worksheet.
    Layout:
      Row 1  – merged title (plain text, NO emojis)
      Row 2  – merged subtitle: period + generated timestamp
      Row 3  – small visual separator
      Row 4  – column header row  (frozen here)
      Row 5+ – data rows with alternating fill
    """
    n_cols = max(len(headers), 1)

    # Row 1 – Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=title)
    t.font      = styles['title_font']
    t.alignment = styles['center']

    # Row 2 – Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    s = ws.cell(
        row=2, column=1,
        value=f"Periodo: {period_label}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    s.font      = styles['subtitle_font']
    s.alignment = styles['center']

    # Row 3 – Visual separator
    ws.row_dimensions[3].height = 6

    # Row 4 – Headers
    for col_idx, header in enumerate(headers, start=1):
        cell            = ws.cell(row=4, column=col_idx, value=str(header))
        cell.fill       = styles['header_fill']
        cell.font       = styles['header_font']
        cell.alignment  = styles['center']
        cell.border     = styles['border']

    # Row 5+ – Data
    for row_idx, row_data in enumerate(rows, start=5):
        use_fill = styles['alt_fill'] if row_idx % 2 == 0 else styles['no_fill']
        for col_idx, value in enumerate(row_data, start=1):
            safe_val       = value if value is not None else ""
            cell           = ws.cell(row=row_idx, column=col_idx, value=safe_val)
            cell.fill      = use_fill
            cell.border    = styles['border']
            cell.alignment = styles['left_vc']

    # Auto-size columns (capped at 50 chars wide)
    for col_idx, header in enumerate(headers, start=1):
        col_letter  = get_column_letter(col_idx)
        col_values  = [
            str(row[col_idx - 1])
            for row in rows
            if row and len(row) >= col_idx and row[col_idx - 1] is not None
        ]
        max_len = max([len(str(header))] + [len(v) for v in col_values] + [0])
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A5"


def _build_response(wb, filename: str) -> HttpResponse:
    """Serialises workbook to bytes and returns an HttpResponse."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        content=buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _csv_fallback(filename: str, headers: list, rows: list) -> HttpResponse:
    """Plain-text CSV fallback when openpyxl is not installed."""
    import csv
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = f'attachment; filename="{filename.replace(".xlsx", ".csv")}"'
    writer = csv.writer(resp)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return resp


# ─── Public API ───────────────────────────────────────────────────────────────

def create_excel_response(
    filename: str,
    sheet_name: str,
    headers: list,
    rows: list,
    title: str,
    period_label: str,
) -> HttpResponse:
    """
    Creates an HttpResponse with a styled single-sheet .xlsx workbook.
    """
    if not OPENPYXL_AVAILABLE:
        return _csv_fallback(filename, headers, rows)

    styles = _make_styles()
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = sheet_name[:31]   # Excel enforces max 31 chars per sheet name
    _write_sheet(ws, headers, rows, title, period_label, styles)
    return _build_response(wb, filename)


def create_multi_sheet_excel_response(
    filename: str,
    sheets_data: list,
) -> HttpResponse:
    """
    Creates an HttpResponse with a styled multi-sheet .xlsx workbook.

    sheets_data – list of dicts, each with keys:
        sheet_name   : str
        headers      : list[str]
        rows         : list[list]
        title        : str
        period_label : str
    """
    if not OPENPYXL_AVAILABLE:
        first = sheets_data[0] if sheets_data else {}
        return _csv_fallback(filename, first.get('headers', []), first.get('rows', []))

    styles = _make_styles()
    wb     = openpyxl.Workbook()

    for idx, sheet_info in enumerate(sheets_data):
        ws       = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_info['sheet_name'][:31]
        _write_sheet(
            ws,
            sheet_info['headers'],
            sheet_info['rows'],
            sheet_info['title'],
            sheet_info['period_label'],
            styles,
        )

    return _build_response(wb, filename)
