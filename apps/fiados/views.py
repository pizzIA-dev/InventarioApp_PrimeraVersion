from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from rest_framework import viewsets, status
import openpyxl
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import ClienteFiado, Fiado, HistorialFiado
from .serializers import ClienteFiadoSerializer, FiadoSerializer, FiadoCreateSerializer

class ClienteFiadoViewSet(viewsets.ModelViewSet):
    queryset = ClienteFiado.objects.all()
    serializer_class = ClienteFiadoSerializer

    def get_queryset(self):
        queryset = self.queryset
        empresa_id = self.request.query_params.get('empresa')
        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        return queryset

    def perform_create(self, serializer):
        # Si viene empresa en la data, usarla, si no usar la de los params
        empresa_id = self.request.data.get('empresa') or self.request.query_params.get('empresa')
        if empresa_id:
            serializer.save(empresa_id=empresa_id)
        else:
            serializer.save()

    # Soft delete
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Verificar si tiene fiados pendientes
        if instance.fiados.filter(estado__in=['PENDIENTE', 'PAGADO_PARCIAL']).exists():
            return Response(
                {"error": "No se puede eliminar un cliente con fiados pendientes."},
                status=status.HTTP_400_BAD_REQUEST
            )
        # Si tiene operaciones antiguas, solo desactivar
        if instance.fiados.exists():
            instance.activo = False
            instance.save()
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Retorna todo el historial de movimientos de todos los fiados del cliente"""
        cliente = self.get_object()
        historial = HistorialFiado.objects.filter(fiado__cliente=cliente).order_by('-fecha')
        
        # Filtros de fecha
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        if fecha_desde:
            historial = historial.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            historial = historial.filter(fecha__date__lte=fecha_hasta)

        # Paginación manual simple
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 15))
        total_count = historial.count()
        
        start = (page - 1) * page_size
        end = start + page_size
        paged_historial = historial[start:end]
        
        data = []
        for h in paged_historial:
            data.append({
                "id": h.id,
                "fiado_id": h.fiado.id,
                "fiado_tipo": h.fiado.tipo,
                "fecha": h.fecha.isoformat(),
                "abono": str(h.abono),
                "saldo_restante": str(h.saldo_restante),
                "estado_nuevo": h.estado_nuevo,
                "notas": h.notas
            })
            
        return Response({
            "count": total_count,
            "results": data,
            "page": page,
            "page_size": page_size
        })

    @action(detail=True, methods=['get'])
    def exportar_historial(self, request, pk=None):
        """Exporta el historial de movimientos de un cliente a Excel"""
        cliente = self.get_object()
        historial = HistorialFiado.objects.filter(fiado__cliente=cliente).order_by('-fecha')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Historial Cliente {cliente.id}"

        # Setup headers
        headers = [
            'ID Fiado', 'Tipo Operación', 'Fecha Movimiento', 
            'Abono Realizado (S/.)', 'Saldo Restante (S/.)', 
            'Estado', 'Notas'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for h in historial:
            ws.append([
                f"#{h.fiado.id}",
                h.fiado.get_tipo_display(),
                h.fecha.strftime("%d/%m/%Y %H:%M:%S") if h.fecha else '',
                float(h.abono),
                float(h.saldo_restante),
                h.estado_nuevo,
                h.notas or ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="historial_cliente_{cliente.id}.xlsx"'
        wb.save(response)
        return response


class FiadoViewSet(viewsets.ModelViewSet):
    queryset = Fiado.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return FiadoCreateSerializer
        return FiadoSerializer

    def get_queryset(self):
        queryset = self.queryset
        empresa_id = self.request.query_params.get('empresa')
        tipo = self.request.query_params.get('tipo')
        estado = self.request.query_params.get('estado')
        cliente = self.request.query_params.get('cliente')

        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if estado:
            queryset = queryset.filter(estado=estado)
        if cliente:
            queryset = queryset.filter(cliente_id=cliente)

        return queryset

    def perform_create(self, serializer):
        # Intentar obtener empresa de la data, o de los params, o del cliente
        empresa_id = self.request.data.get('empresa') or self.request.query_params.get('empresa')
        
        if not empresa_id:
            # Buscar la empresa del cliente seleccionado
            cliente_id = self.request.data.get('cliente')
            if cliente_id:
                try:
                    cliente = ClienteFiado.objects.get(id=cliente_id)
                    empresa_id = cliente.empresa_id
                except ClienteFiado.DoesNotExist:
                    pass
        
        if empresa_id:
            serializer.save(empresa_id=empresa_id)
        else:
            serializer.save()
        
    @action(detail=True, methods=['post'])
    def abonar(self, request, pk=None):
        """Registra un pago parcial o total en el fiado"""
        fiado = self.get_object()
        
        try:
            monto = Decimal(str(request.data.get('monto', 0)))
            notas = request.data.get('notas', '')
        except (ValueError, InvalidOperation):
            return Response({"error": "El monto debe ser un valor numérico válido"}, status=status.HTTP_400_BAD_REQUEST)
            
        if monto <= 0:
            return Response({"error": "El abono debe ser mayor a 0"}, status=status.HTTP_400_BAD_REQUEST)
            
        if monto > fiado.saldo_pendiente:
            return Response({"error": "El abono no puede superar el saldo pendiente"}, status=status.HTTP_400_BAD_REQUEST)
            
        # Aplicar el abono
        fiado.saldo_pendiente -= monto
        
        # Guardar (el save ya calculará el estado automáticamente)
        fiado.save()
        
        # Registrar en el historial manualmente para que registre el abono per-se
        HistorialFiado.objects.create(
            fiado=fiado,
            abono=monto,
            saldo_restante=fiado.saldo_pendiente,
            estado_nuevo=fiado.estado,
            notas=notas or f"Abono de S/ {monto:.2f} registrado."
        )
        
        return Response(FiadoSerializer(fiado).data)

    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Cancela el fiado y revierte el stock retirado de productos"""
        fiado = self.get_object()
        
        if fiado.estado == 'CANCELADO':
            return Response({"error": "El fiado ya se encuentra cancelado."}, status=status.HTTP_400_BAD_REQUEST)
            
        if fiado.venta_ref or fiado.venta_servicio_ref:
            return Response({"error": "No se puede cancelar un fiado que ya fue formalizado en Venta."}, status=status.HTTP_400_BAD_REQUEST)
            
        fiado.estado = 'CANCELADO'
        fiado.save()
        
        # Devolver stock
        fiado.revertir_stock()
        
        HistorialFiado.objects.create(
            fiado=fiado,
            abono=0,
            saldo_restante=fiado.saldo_pendiente,
            estado_nuevo='CANCELADO',
            notas="Fiado CANCELADO manualmente. Stock revertido."
        )
        
        return Response(FiadoSerializer(fiado).data)

    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Retorna el historial de movimientos de un fiado específico"""
        fiado = self.get_object()
        historial = fiado.historial.all().order_by('-fecha')
        
        # Filtros de fecha
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        if fecha_desde:
            historial = historial.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            historial = historial.filter(fecha__date__lte=fecha_hasta)

        # Paginación
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 15))
        total_count = historial.count()

        start = (page - 1) * page_size
        end = start + page_size
        paged_historial = historial[start:end]
        
        data = []
        for h in paged_historial:
            data.append({
                "id": h.id,
                "fecha": h.fecha.isoformat(),
                "abono": str(h.abono),
                "saldo_restante": str(h.saldo_restante),
                "estado_nuevo": h.estado_nuevo,
                "notas": h.notas
            })
            
        return Response({
            "count": total_count,
            "results": data,
            "page": page,
            "page_size": page_size
        })

    @action(detail=True, methods=['get'])
    def exportar_historial(self, request, pk=None):
        """Exporta el historial de un fiado a Excel"""
        fiado = self.get_object()
        historial = fiado.historial.all().order_by('-fecha')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Historial Fiado {fiado.id}"

        headers = [
            'Fecha Movimiento', 'Abono Registrado (S/.)', 
            'Saldo Restante (S/.)', 'Estado Resultante', 'Notas'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for h in historial:
            ws.append([
                h.fecha.strftime("%d/%m/%Y %H:%M:%S") if h.fecha else '',
                float(h.abono),
                float(h.saldo_restante),
                h.estado_nuevo,
                h.notas or ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="historial_fiado_{fiado.id}.xlsx"'
        wb.save(response)
        return response
